# -*- coding: utf-8 -*-
"""
Módulo de importación de listas de asistencia para constancias de webinar.

Soporta varios formatos de reporte (Microsoft Teams, WebEx/GoToWebinar, Zoom y
Excel genérico) y permite completar los correos faltantes cruzando contra un
"padrón" de inscritos (el Excel de capacitaciones).

Uso típico:
    asistentes, formato = parsear_asistencia(archivo)
    padron = cargar_padron(archivo_padron)          # opcional
    asistentes, pendientes = completar_correos(asistentes, padron)
"""

import csv
import io
import re
import unicodedata

import openpyxl

# --- Títulos profesionales que se ignoran al comparar nombres ---
TITULOS = {
    'q', 'qfb', 'qbp', 'qc', 'qa', 'quim', 'dr', 'dra', 'lic', 'ing',
    'mtro', 'mtra', 'mc', 'm', 'biol', 'tm', 'ct', 'md', 'phd', 'sr', 'sra',
}

# =============================================================================
#  NORMALIZACIÓN Y COMPARACIÓN DE NOMBRES
# =============================================================================

def normalizar_nombre(texto):
    """Minúsculas, sin acentos, sin paréntesis, sin puntuación.
    Convierte 'Apellido, Nombre' -> 'Nombre Apellido'."""
    if not isinstance(texto, str):
        return ''
    t = re.sub(r'\(.*?\)', ' ', texto)          # quita "(No comprobado)", etc.
    if ',' in t:
        partes = [p.strip() for p in t.split(',')]
        if len(partes) == 2:
            t = partes[1] + ' ' + partes[0]
    t = unicodedata.normalize('NFD', t)
    t = ''.join(c for c in t if unicodedata.category(c) != 'Mn')
    t = re.sub(r'[^a-z\s]', ' ', t.lower())
    return re.sub(r'\s+', ' ', t).strip()


def tokens_nombre(texto):
    """Palabras significativas del nombre (sin títulos ni iniciales sueltas)."""
    return [p for p in normalizar_nombre(texto).split()
            if p not in TITULOS and len(p) > 1]


def parsear_duracion(texto):
    """Convierte duraciones a minutos enteros.
    Acepta '1 h 48 min 15s', '58 min 6s', '45 min', '90', '1:30:00'."""
    if texto is None:
        return 0
    if isinstance(texto, (int, float)):
        return int(texto)
    t = str(texto).strip().lower()
    if not t:
        return 0

    # Formato hh:mm:ss
    if ':' in t and re.match(r'^\d+:\d{1,2}(:\d{1,2})?$', t):
        partes = [int(p) for p in t.split(':')]
        if len(partes) == 3:
            return partes[0] * 60 + partes[1]
        return partes[0] * 60 + partes[1] if partes[0] > 12 else partes[0] * 60 + partes[1]

    minutos = 0
    encontrado = False
    h = re.search(r'(\d+)\s*h', t)
    if h:
        minutos += int(h.group(1)) * 60
        encontrado = True
    m = re.search(r'(\d+)\s*min', t)
    if m:
        minutos += int(m.group(1))
        encontrado = True
    if not encontrado:
        num = re.match(r'^(\d+)', t)
        if num:
            minutos = int(num.group(1))
    return minutos


def _es_correo(valor):
    return bool(isinstance(valor, str) and re.match(r'^[^@\s]+@[^@\s]+\.[^@\s]+$', valor.strip()))


# =============================================================================
#  DETECCIÓN DE COLUMNAS POR ENCABEZADO
# =============================================================================

ALIAS = {
    'nombre':      ['nombre completo', 'nombre', 'name', 'full name', 'attendee', 'participante', 'asistente'],
    'primer':      ['first name', 'nombre(s)', 'nombres'],
    'apellido':    ['last name', 'apellido', 'apellidos'],
    'email':       ['correo electronico', 'correo', 'email', 'e-mail', 'mail'],
    'institucion': ['institucion', 'empresa', 'company', 'organizacion', 'organization'],
    'duracion':    ['duracion de la reunion', 'duracion', 'tiempo de asistencia',
                    'attendance duration', 'time in session', 'duration', 'minutos'],
}


def _mapear_columnas(encabezados):
    """Devuelve {campo: indice} a partir de una fila de encabezados."""
    limpios = [normalizar_nombre(str(h or '')) for h in encabezados]
    mapa = {}
    for campo, opciones in ALIAS.items():
        for opcion in opciones:
            objetivo = normalizar_nombre(opcion)
            for i, h in enumerate(limpios):
                if h == objetivo and campo not in mapa:
                    mapa[campo] = i
                    break
            if campo in mapa:
                break
    # Segunda pasada: coincidencia parcial
    for campo, opciones in ALIAS.items():
        if campo in mapa:
            continue
        for opcion in opciones:
            objetivo = normalizar_nombre(opcion)
            for i, h in enumerate(limpios):
                if objetivo and objetivo in h and i not in mapa.values():
                    mapa[campo] = i
                    break
            if campo in mapa:
                break
    return mapa


def _fila_de_encabezado(filas, limite=30):
    """Busca la fila que funciona como encabezado.

    Exige una columna de nombre Y al menos otra columna reconocible, para no
    confundir títulos de sección (p. ej. '2. Participantes') con encabezados.
    """
    mejor = (None, {}, 0)
    for i, fila in enumerate(filas[:limite]):
        if not fila or sum(1 for c in fila if c is not None and str(c).strip()) < 2:
            continue
        mapa = _mapear_columnas(fila)
        tiene_nombre = 'nombre' in mapa or ('primer' in mapa and 'apellido' in mapa)
        if not tiene_nombre or len(mapa) < 2:
            continue
        if len(mapa) > mejor[2]:
            mejor = (i, mapa, len(mapa))
        if 'email' in mapa and 'duracion' in mapa:
            return i, mapa
    return mejor[0], mejor[1]


# =============================================================================
#  LECTURA DE ARCHIVOS
# =============================================================================

def _decodificar(datos):
    for enc in ('utf-16', 'utf-8-sig', 'utf-8', 'latin-1'):
        try:
            texto = datos.decode(enc)
            if texto.count('\x00') > len(texto) * 0.1:
                continue
            return texto
        except (UnicodeDecodeError, UnicodeError):
            continue
    return None


def _filas_desde_texto(texto):
    muestra = texto[:5000]
    delim = '\t' if muestra.count('\t') >= muestra.count(',') else ','
    return [f for f in csv.reader(io.StringIO(texto), delimiter=delim)]


def _filas_desde_excel(datos, hoja=None):
    wb = openpyxl.load_workbook(io.BytesIO(datos), data_only=True, read_only=True)
    ws = wb[hoja] if hoja and hoja in wb.sheetnames else wb.active
    return [list(f) for f in ws.iter_rows(values_only=True)]


# =============================================================================
#  PARSEO PRINCIPAL DE ASISTENCIA
# =============================================================================

def parsear_asistencia(archivo, nombre_archivo=None, hoja=None):
    """Lee un reporte de asistencia y devuelve (asistentes, formato_detectado).

    Cada asistente: {'nombre_completo', 'email', 'institucion', 'duracion_total'}
    Las sesiones repetidas de la misma persona se suman.
    """
    datos = archivo.read() if hasattr(archivo, 'read') else archivo
    nombre_archivo = (nombre_archivo or getattr(archivo, 'name', '') or '').lower()

    if nombre_archivo.endswith(('.xlsx', '.xlsm', '.xls')):
        filas = _filas_desde_excel(datos, hoja)
        formato = 'excel'
    else:
        texto = _decodificar(datos)
        if texto is None:
            raise ValueError("No se pudo leer el archivo. Guárdalo como CSV UTF-8 o Excel.")
        filas = _filas_desde_texto(texto)
        formato = 'csv'

    # Teams divide el reporte en secciones numeradas; nos quedamos con
    # "2. Participantes" y cortamos antes de "3. Actividades".
    inicio, fin = 0, len(filas)
    for i, fila in enumerate(filas):
        primera = str(fila[0]).strip().lower() if fila and fila[0] else ''
        if re.match(r'^2\.\s', primera) and 'participante' in primera:
            inicio, formato = i, 'teams'
        elif re.match(r'^[3-9]\.\s', primera) and inicio:
            fin = i
            break
    filas = filas[inicio:fin]

    idx_encabezado, mapa = _fila_de_encabezado(filas)
    if idx_encabezado is None:
        raise ValueError(
            "No se encontró una columna de nombres. Verifica que el archivo "
            "tenga encabezados como 'Nombre' y 'Correo electrónico'."
        )

    acumulado = {}
    sin_correo = []

    for fila in filas[idx_encabezado + 1:]:
        if not fila or all(c is None or str(c).strip() == '' for c in fila):
            continue

        def celda(campo):
            i = mapa.get(campo)
            if i is None or i >= len(fila) or fila[i] is None:
                return ''
            return str(fila[i]).strip()

        if 'nombre' in mapa:
            nombre = celda('nombre')
        else:
            nombre = (celda('primer') + ' ' + celda('apellido')).strip()

        nombre = re.sub(r'\s*\(no comprobado\)\s*', '', nombre, flags=re.I).strip()
        if not nombre or normalizar_nombre(nombre) in ('nombre', 'name'):
            continue

        email = celda('email').lower()
        if not _es_correo(email):
            email = ''

        registro = {
            'nombre_completo': nombre,
            'email': email,
            'institucion': celda('institucion') or 'N/A',
            'duracion_total': parsear_duracion(celda('duracion')),
        }

        clave = email or ('#sin#' + normalizar_nombre(nombre))
        if clave in acumulado:
            acumulado[clave]['duracion_total'] += registro['duracion_total']
            if not acumulado[clave]['email'] and email:
                acumulado[clave]['email'] = email
        else:
            acumulado[clave] = registro

    asistentes = list(acumulado.values())
    return asistentes, formato


# =============================================================================
#  PADRÓN DE INSCRITOS (Excel de capacitaciones)
# =============================================================================

def cargar_padron(archivo, hojas=None):
    """Lee el Excel de inscritos y devuelve [{'nombre','email','institucion','hoja'}].

    Recorre TODAS las hojas por defecto; el encabezado puede estar en
    cualquier fila (los archivos suelen traer un título arriba).
    """
    datos = archivo.read() if hasattr(archivo, 'read') else archivo
    wb = openpyxl.load_workbook(io.BytesIO(datos), data_only=True, read_only=True)

    padron = []
    for nombre_hoja in (hojas or wb.sheetnames):
        if nombre_hoja not in wb.sheetnames:
            continue
        filas = [list(f) for f in wb[nombre_hoja].iter_rows(values_only=True)]
        idx, mapa = _fila_de_encabezado(filas)
        if idx is None or 'email' not in mapa:
            continue

        for fila in filas[idx + 1:]:
            if not fila:
                continue

            def celda(campo):
                i = mapa.get(campo)
                if i is None or i >= len(fila) or fila[i] is None:
                    return ''
                return str(fila[i]).strip()

            nombre = celda('nombre') or (celda('primer') + ' ' + celda('apellido')).strip()
            email = celda('email')
            if not nombre or not _es_correo(email):
                continue

            padron.append({
                'nombre': nombre,
                'email': email,
                'institucion': celda('institucion') or 'N/A',
                'hoja': nombre_hoja,
            })
    return padron


# =============================================================================
#  CRUCE: COMPLETAR CORREOS FALTANTES
# =============================================================================

def buscar_en_padron(nombre, padron):
    """Busca una persona en el padrón por nombre.
    Devuelve (registro, estado) donde estado es 'ok', 'ambiguo',
    'no_inscrito' o 'nombre_muy_corto'."""
    t = set(tokens_nombre(nombre))
    if len(t) < 2:
        return None, 'nombre_muy_corto'

    candidatos = []
    for reg in padron:
        p = set(tokens_nombre(reg['nombre']))
        comunes = t & p
        if len(comunes) >= 2:
            candidatos.append((len(comunes), t <= p, reg))

    if not candidatos:
        return None, 'no_inscrito'

    # Varias filas pueden ser la misma persona duplicada en el padrón.
    correos = {c[2]['email'].strip().lower() for c in candidatos}
    if len(correos) == 1:
        mejor = max(candidatos, key=lambda x: (x[1], x[0]))
        return mejor[2], 'ok'

    candidatos.sort(key=lambda x: (x[1], x[0]), reverse=True)
    if candidatos[0][0] == candidatos[1][0] and candidatos[0][1] == candidatos[1][1]:
        return [c[2] for c in candidatos[:4]], 'ambiguo'
    return candidatos[0][2], 'ok'


def completar_correos(asistentes, padron):
    """Rellena los correos faltantes usando el padrón.

    Devuelve (asistentes, pendientes). Los 'pendientes' NUNCA se adivinan:
    se devuelven para que una persona los complete o descarte a mano.
    """
    pendientes = []
    if not padron:
        pendientes = [a for a in asistentes if not a['email']]
        for a in pendientes:
            a['estado_correo'] = 'sin_padron'
        return asistentes, pendientes

    for a in asistentes:
        if a['email']:
            a['estado_correo'] = 'original'
            continue

        resultado, estado = buscar_en_padron(a['nombre_completo'], padron)
        if estado == 'ok':
            a['email'] = resultado['email'].lower()
            a['estado_correo'] = 'recuperado'
            a['origen_correo'] = resultado['nombre']
            if a['institucion'] in ('', 'N/A'):
                a['institucion'] = resultado['institucion']
        else:
            a['estado_correo'] = estado
            a['sugerencias'] = (
                [r['email'] for r in resultado] if estado == 'ambiguo' and resultado else []
            )
            pendientes.append(a)

    # Al recuperar correos pueden aparecer duplicados: los fusionamos.
    fusionado = {}
    for a in asistentes:
        clave = a['email'] or ('#sin#' + normalizar_nombre(a['nombre_completo']))
        if clave in fusionado and a['email']:
            fusionado[clave]['duracion_total'] += a['duracion_total']
        else:
            fusionado[clave] = a

    return list(fusionado.values()), pendientes


# =============================================================================
#  LIBRO DE CAPACITACIONES: UNA SESIÓN POR HOJA, CON CALIFICACIONES
# =============================================================================
#
#  Estructura esperada de cada hoja:
#     fila 0    -> título: "Técnica de ELISA, 11 de agosto, remoto, 11:00 a 12:30"
#     fila 1    -> encabezado de INSCRITOS (Número, Nombre, Institución, Correo, ...)
#     filas...  -> inscritos
#     (fila en blanco)
#     filas...  -> ASISTENTES, con una columna de Calificación
#     footer    -> "Porcentaje de asistencia", "Asistentes finales", etc.
#
#  El encabezado del bloque de asistentes es inconsistente entre hojas
#  (a veces completo, a veces sólo "Calificación", a veces ausente),
#  por eso el bloque se detecta por contenido y no por posición.
# =============================================================================

MESES = {
    'enero': 1, 'febrero': 2, 'marzo': 3, 'abril': 4, 'mayo': 5, 'junio': 6,
    'julio': 7, 'agosto': 8, 'septiembre': 9, 'setiembre': 9, 'octubre': 10,
    'noviembre': 11, 'diciembre': 12,
}

FOOTER = ('porcentaje de', 'asistentes finales', 'laboratorios participantes')


def _a_numero(valor):
    """Devuelve float si la celda es una calificación (0-100); si no, None.
    Rechaza teléfonos ('33 1047 6002', '4432424008') y textos ('NA')."""
    if valor is None:
        return None
    if isinstance(valor, bool):
        return None
    if isinstance(valor, (int, float)):
        n = float(valor)
    else:
        t = str(valor).strip().replace('%', '')
        if not re.match(r'^\d{1,3}(\.\d+)?$', t):
            return None
        n = float(t)
    return n if 0 <= n <= 100 else None


def interpretar_titulo(titulo, anio=None):
    """Extrae curso, fecha y duración del título de la hoja.
    Ej: 'Técnica de ELISA, 11 de agosto, remoto, 11:00 a 12:30'."""
    import datetime

    datos = {'curso': '', 'fecha': None, 'duracion_horas': None, 'modalidad': ''}
    if not titulo:
        return datos

    partes = [p.strip() for p in str(titulo).split(',') if p.strip()]
    if partes:
        datos['curso'] = partes[0].rstrip(')').strip()

    texto = str(titulo).lower()

    m = re.search(r'(\d{1,2})\s+de\s+([a-záéíóú]+)', texto)
    if m:
        dia = int(m.group(1))
        mes = MESES.get(unicodedata.normalize('NFD', m.group(2))
                        .encode('ascii', 'ignore').decode())
        if mes:
            año = anio or datetime.date.today().year
            a = re.search(r'\b(20\d{2})\b', texto)
            if a:
                año = int(a.group(1))
            try:
                datos['fecha'] = datetime.date(año, mes, dia)
            except ValueError:
                pass

    h = re.search(r'(\d{1,2}):(\d{2})\s*(?:a|-|–)\s*(\d{1,2}):(\d{2})', texto)
    if h:
        ini = int(h.group(1)) * 60 + int(h.group(2))
        fin = int(h.group(3)) * 60 + int(h.group(4))
        if fin > ini:
            datos['duracion_horas'] = round((fin - ini) / 60.0, 1)

    for palabra in ('remoto', 'online', 'presencial', 'híbrido', 'hibrido'):
        if palabra in texto:
            datos['modalidad'] = palabra
            break

    return datos


def _bloque_calificaciones(filas, inicio, mapa):
    """Localiza la columna de calificación y las filas de asistentes.
    Devuelve (col_calificacion, [filas])."""
    usadas = {mapa.get(c) for c in ('nombre', 'email', 'institucion')}
    usadas.discard(None)

    candidatas = {}
    filas_bloque = []

    for fila in filas[inicio:]:
        if not fila:
            continue
        etiqueta = str(fila[1]).strip().lower() if len(fila) > 1 and fila[1] else ''
        if any(etiqueta.startswith(f) for f in FOOTER):
            break

        idx_nombre = mapa.get('nombre', 1)
        nombre = (str(fila[idx_nombre]).strip()
                  if idx_nombre < len(fila) and fila[idx_nombre] else '')
        if not nombre or normalizar_nombre(nombre) in ('nombre', 'name'):
            continue

        filas_bloque.append(fila)
        for j, celda in enumerate(fila):
            if j in usadas or j == 0:
                continue
            if _a_numero(celda) is not None:
                candidatas[j] = candidatas.get(j, 0) + 1

    if not candidatas:
        return None, filas_bloque
    col = max(candidatas, key=lambda k: candidatas[k])
    return col, filas_bloque


def analizar_libro(archivo, anio=None, calificacion_minima=80):
    """Analiza un libro de capacitaciones y devuelve una sesión por hoja.

    Cada sesión:
      {'hoja','titulo','curso','fecha','duracion_horas','modalidad',
       'inscritos','asistentes':[...],'aprobados','reprobados',
       'sin_correo','estado'}

    'estado' es 'listo' (hay calificaciones) o 'sin_calificar'.
    """
    datos = archivo.read() if hasattr(archivo, 'read') else archivo
    wb = openpyxl.load_workbook(io.BytesIO(datos), data_only=True, read_only=True)

    sesiones = []
    for nombre_hoja in wb.sheetnames:
        filas = [list(f) for f in wb[nombre_hoja].iter_rows(values_only=True)]
        if not filas:
            continue

        titulo = str(filas[0][0]).strip() if filas[0] and filas[0][0] else nombre_hoja
        meta = interpretar_titulo(titulo, anio)

        idx, mapa = _fila_de_encabezado(filas)
        if idx is None or 'nombre' not in mapa:
            continue

        # Fin del bloque de inscritos: primera fila sin nombre.
        i = idx + 1
        inscritos = 0
        while i < len(filas):
            fila = filas[i]
            col = mapa['nombre']
            valor = (str(fila[col]).strip()
                     if fila and col < len(fila) and fila[col] else '')
            if not valor:
                break
            inscritos += 1
            i += 1

        col_calif, filas_asistentes = _bloque_calificaciones(filas, i, mapa)

        asistentes = []
        for fila in filas_asistentes:
            def celda(campo):
                j = mapa.get(campo)
                if j is None or j >= len(fila) or fila[j] is None:
                    return ''
                return str(fila[j]).strip()

            nombre = celda('nombre')
            email = celda('email').lower()
            if not _es_correo(email):
                email = ''

            calif = None
            if col_calif is not None and col_calif < len(fila):
                calif = _a_numero(fila[col_calif])

            asistentes.append({
                'nombre_completo': nombre,
                'email': email,
                'institucion': celda('institucion') or 'N/A',
                'calificacion': calif,
                'aprueba': calif is not None and calif >= calificacion_minima,
            })

        con_calif = [a for a in asistentes if a['calificacion'] is not None]
        aprobados = [a for a in asistentes if a['aprueba'] and a['email']]

        sesiones.append({
            'hoja': nombre_hoja,
            'titulo': titulo,
            'curso': meta['curso'] or nombre_hoja,
            'fecha': meta['fecha'],
            'duracion_horas': meta['duracion_horas'],
            'modalidad': meta['modalidad'],
            'inscritos': inscritos,
            'asistentes': asistentes,
            'con_calificacion': len(con_calif),
            'aprobados': len(aprobados),
            'reprobados': len([a for a in asistentes
                               if a['calificacion'] is not None and not a['aprueba']]),
            'sin_correo': len([a for a in asistentes if a['aprueba'] and not a['email']]),
            'estado': 'listo' if con_calif else 'sin_calificar',
        })

    return sesiones